from openpyxl import load_workbook


def read_excel(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    data = []

    for row in sheet.iter_rows():
        row_list = []
        for cell in row:
            cell_value = str(cell.value)         # Cell Value
            cell_font = cell.font                # Cell Font
            cell_border = cell.border            # Cell Boarder
            cell_parent = cell.parent            # Cell Parent
            cell_coordinate = cell.coordinate    # Cell Coordinate
            
            row_list.append([cell_coordinate, cell_value, cell_font, cell_border, cell_parent])
        # row_data = [(str(cell.value), cell.font) for cell in row]
        data.append(row_list)
    return data

# Main execution
file_path = "C:\\Users\\kyleg\\Downloads\\Schedules_FtWayneBoiler.xlsx"
sheet_name = "M-SCH-AirDev-Supply (TIL)"
view_name = "M-SCH-AirDev-Supply (TIL) 001"

schedule_data = read_excel(file_path, sheet_name)
if not schedule_data:
    raise Exception("Excel data is empty or could not be read.")

